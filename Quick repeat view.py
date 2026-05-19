import os
import glob
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG ───────────────────────────────────────────────────────────────────
STOCK_FILE   = r'C:\Users\User\Desktop\daily stock 17_05_2026.xlsx'
SALES_FOLDER = r'D:\mahmoud_analysis_dashboard_folder\spikes\sales_files'
OUTPUT_FILE  = r'C:\Users\User\Desktop\section_sales_output.xlsx'
# ──────────────────────────────────────────────────────────────────────────────


def load_stock(sections: list[str]) -> pd.DataFrame:
    """Load stock file and filter by selected sections."""
    df = pd.read_excel(STOCK_FILE, dtype={'ARTCODE': str, 'ARTSEC': str})
    # Find stock column regardless of surrounding spaces
    stock_col = next((c for c in df.columns if 'STOCK WITHOUT SMALL ROLL' in c.upper()), None)
    if stock_col is None:
        raise ValueError(f"Could not find 'STOCK WITHOUT SMALL ROLL' column. Available: {list(df.columns)}")
    df = df[df['ARTSEC'].isin(sections)][['ARTSEC', 'ARTCODE', stock_col]].drop_duplicates()
    df = df.rename(columns={stock_col: 'stock'})
    return df


def load_sales_file(path: str) -> pd.DataFrame:
    """Load one sales xlsx and return SKU → TOTAL mapping."""
    df = pd.read_excel(path, dtype={'SKU': str})
    # Keep only SKU and TOTAL
    df = df[['SKU', 'CATEGORY', 'TOTAL']].copy()
    df.columns = ['SKU', 'CATEGORY', 'TOTAL']
    return df


def detect_year(filename: str) -> str:
    """Extract year from filename like 2024sales.xlsx → '2024'."""
    basename = os.path.basename(filename)
    for year in ['2024', '2025', '2026']:
        if year in basename:
            return year
    return 'unknown'


def run_analysis(sections: list[str]) -> str:
    # 1. Load stock → get SKUs per section
    stock_df = load_stock(sections)

    # 2. Load all sales files from the folder
    sales_files = glob.glob(os.path.join(SALES_FOLDER, '*.xlsx'))
    sales_by_year = {}

    for path in sales_files:
        year = detect_year(path)
        sales_df = load_sales_file(path)
        sales_by_year[year] = sales_df

    # 3. Merge stock with each year's sales
    result = stock_df.copy()

    for year in ['2024', '2025', '2026']:
        if year in sales_by_year:
            sales = sales_by_year[year][['SKU', 'TOTAL']].rename(columns={'TOTAL': f'sales_{year}'})
            result = result.merge(sales, left_on='ARTCODE', right_on='SKU', how='left').drop(columns='SKU')
        else:
            result[f'sales_{year}'] = None

    # Add CATEGORY from any available sales file
    if sales_by_year:
        any_sales = next(iter(sales_by_year.values()))[['SKU', 'CATEGORY']]
        result = result.merge(any_sales, left_on='ARTCODE', right_on='SKU', how='left').drop(columns='SKU')

    # 4. Reorder columns
    cols = ['ARTSEC', 'ARTCODE', 'CATEGORY', 'stock', 'sales_2024', 'sales_2025', 'sales_2026']
    cols = [c for c in cols if c in result.columns]
    result = result[cols]
    result.columns = ['section', 'sku', 'category', 'stock', 'sales_2024', 'sales_2025', 'sales_2026'][:len(cols)]

    # 5. Write to Excel with styling
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Analysis'

    # Header style
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial')

    for col_idx, col_name in enumerate(result.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows — alternate row shading per section
    section_colors = {}
    color_pool = ['EBF3FB', 'FDFEFE', 'EAF5EA', 'FEF9E7', 'FDEDEC']
    color_idx = 0

    for row_idx, row in result.iterrows():
        sec = row.get('section', '')
        if sec not in section_colors:
            section_colors[sec] = color_pool[color_idx % len(color_pool)]
            color_idx += 1
        fill_color = section_colors[sec]

        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.fill = PatternFill('solid', start_color=fill_color)
            cell.font = Font(name='Arial', size=10)
            if col_idx >= 4:  # stock + sales columns → number format
                cell.number_format = '#,##0'

    # Column widths
    widths = [14, 14, 20, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


# ─── UI ───────────────────────────────────────────────────────────────────────

def build_ui():
    root = tk.Tk()
    root.title('Section Sales Analysis')
    root.geometry('480x400')
    root.resizable(False, False)
    root.configure(bg='#1F4E79')

    tk.Label(root, text='Section Sales Analysis', font=('Arial', 16, 'bold'),
             bg='#1F4E79', fg='white').pack(pady=(20, 4))

    tk.Label(root, text='Enter sections (one per line):', font=('Arial', 10),
             bg='#1F4E79', fg='#AED6F1').pack()

    text_box = tk.Text(root, height=10, width=40, font=('Consolas', 11),
                       bd=0, relief='flat', bg='#EBF5FB', fg='#1A252F')
    text_box.pack(pady=10, padx=30)

    status_var = tk.StringVar(value='')
    status_lbl = tk.Label(root, textvariable=status_var, font=('Arial', 9),
                          bg='#1F4E79', fg='#A9DFBF', wraplength=420)
    status_lbl.pack(pady=4)

    def on_run():
        raw = text_box.get('1.0', tk.END).strip()
        sections = [s.strip() for s in raw.splitlines() if s.strip()]
        if not sections:
            messagebox.showwarning('Input needed', 'Please enter at least one section.')
            return
        status_var.set('⏳ Processing...')
        root.update()
        try:
            out = run_analysis(sections)
            status_var.set(f'✅ Done! Saved to:\n{out}')
        except Exception as e:
            status_var.set(f'❌ Error: {e}')

    btn = tk.Button(root, text='▶  Run Analysis', font=('Arial', 11, 'bold'),
                    bg='#27AE60', fg='white', relief='flat', padx=20, pady=8,
                    cursor='hand2', command=on_run)
    btn.pack(pady=6)

    root.mainloop()


if __name__ == '__main__':
    build_ui()
