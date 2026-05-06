"""
╔══════════════════════════════════════════════════════════════════════════╗
║                     SPIKE ALARM — Daily Stock Monitor                    ║
║  الإصدار النهائي المستقر — مع تطبيع SKU وقراءة مرنة لـ COGS             ║
║  وجمع الـ Outstanding الثلاثة في عمود واحد                               ║
║  تم إصلاح خطأ 'int' object has no attribute 'strip'                     ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import glob
import re
import traceback
from datetime import datetime

import pandas as pd
import win32com.client
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
#  الإعدادات — عدّلها حسب بيئتك
# ═══════════════════════════════════════════════════════════════════════════

EXCEL_FOLDER = os.path.dirname(os.path.abspath(__file__))
REPORT_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "SpikeReports")
EMAIL_TO = "s.inventory.planner@alguthmi.com;consultant@alguthmi.com"

SPIKE_THRESHOLD = -30.0
MIN_ABSOLUTE_DROP = 300
STOCK_COLUMN = "STOCK WITHOUT SMALL ROLL"
SKU_COLUMN = "ARTCODE"

# أعمدة الـ Outstanding التي سيتم جمعها
OUTSTANDING_COLS = ["OUTSTANDING", "OUTSTANDING SP", "OUTSTANDING N"]
OTHER_EXTRA_COLUMNS = ["PR QTY"]  # أعمدة إضافية أخرى لا تجمع

SALES_FOLDER = os.path.join(EXCEL_FOLDER, "sales_files")
SALES_FILES = {2025: "2025sales.xlsx", 2026: "2026sales.xlsx"}
SALES_EXTRA_COLUMNS = ["CATEGORY", "Section_Ro_Name"]

# مسار ملف COGS
COGS_FILE_PATH = r"D:\COGS-FILES\TOTAL_COGS.xlsx"

# ═══════════════════════════════════════════════════════════════════════════
#  دوال مساعدة
# ═══════════════════════════════════════════════════════════════════════════

def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}]  {msg}")

def clean_column_names(df):
    """تحويل جميع أسماء الأعمدة إلى نصوص، إزالة المسافات الزائدة، واستبدال الأرقام بنصوص"""
    df.columns = [str(col).strip() for col in df.columns]
    return df

def find_column_flexible(df, possible_names):
    """البحث عن عمود في DataFrame يحتوي أي من الأسماء الممكنة (غير حساس لحالة الأحرف)"""
    # تأكد من أن أسماء الأعمدة كلها نصوص
    df_cols_lower = {str(col).lower().strip(): col for col in df.columns}
    for name in possible_names:
        name_lower = name.lower().strip()
        if name_lower in df_cols_lower:
            return df_cols_lower[name_lower]
    return None

def extract_date_label(filepath: str) -> str:
    name = os.path.basename(filepath)
    m = re.search(r'(\d{2})[_\-](\d{2})[_\-](\d{4})', name)
    if m:
        d, mo, y = m.groups()
        return f"{d}/{mo}/{y}"
    return name

def find_latest_two_files(folder: str):
    pattern = os.path.join(folder, "daily stock *.xlsx")
    files = glob.glob(pattern)
    if len(files) < 2:
        raise FileNotFoundError(
            f"❌ محتاج على الأقل ملفين بالاسم 'daily stock *.xlsx' في:\n   {folder}\n"
            f"الموجودة: {[os.path.basename(f) for f in files]}"
        )
    files.sort(key=os.path.getmtime)
    older, newer = files[-2], files[-1]
    log(f"الملف الأقدم  ← {os.path.basename(older)}")
    log(f"الملف الأحدث ← {os.path.basename(newer)}")
    return older, newer

def normalize_sku_value(val):
    """تطبيع قيمة SKU واحدة: تحويل إلى نص، إزالة المسافات، إزالة .0 إذا كان الرقم صحيحاً"""
    if pd.isna(val):
        return ""
    # تحويل إلى نص
    s = str(val).strip()
    # إذا كان النص ينتهي بـ .0 مثلاً "123.0" نحوله إلى "123"
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    return s

def normalize_sku(df, column):
    """تطبيع عمود SKU بالكامل (تحويل إلى نص موحد)"""
    if column in df.columns:
        df[column] = df[column].apply(normalize_sku_value)
    return df

# ═══════════════════════════════════════════════════════════════════════════
#  تحميل البيانات الخارجية
# ═══════════════════════════════════════════════════════════════════════════

def load_sales_data():
    try:
        if not os.path.exists(SALES_FOLDER):
            log(f"⚠️ مجلد المبيعات غير موجود: {SALES_FOLDER}")
            return None

        sales_data = {}
        extra_info = None

        for year, filename in SALES_FILES.items():
            filepath = os.path.join(SALES_FOLDER, filename)
            if not os.path.exists(filepath):
                log(f"⚠️ {filename} غير موجود")
                continue

            df = pd.read_excel(filepath)
            # تنظيف أسماء الأعمدة (تحويل الكل إلى نصوص و strip)
            df = clean_column_names(df)
            log(f"📄 أعمدة {filename}: {list(df.columns)}")

            sku_col = find_column_flexible(df, [SKU_COLUMN, "SKU", "ITEM", "ARTCODE", "CODE"])
            if not sku_col:
                log(f"❌ لم يتم العثور على عمود SKU في {filename}")
                continue

            total_col = find_column_flexible(df, ["TOTAL", "GRAND TOTAL", "مجموع", "إجمالي", "SUM"])
            if not total_col:
                log(f"❌ لم يتم العثور على عمود Total في {filename}")
                continue

            # تطبيع SKU
            df = normalize_sku(df, sku_col)

            temp = df[[sku_col, total_col]].copy()
            temp = temp.rename(columns={sku_col: SKU_COLUMN, total_col: f"Sales_{year}"})
            temp[f"Sales_{year}"] = pd.to_numeric(temp[f"Sales_{year}"], errors="coerce")
            sales_data[year] = temp

            if year == 2025 and extra_info is None:
                extra_found = []
                for col_name in SALES_EXTRA_COLUMNS:
                    if col_name in df.columns:
                        extra_found.append(col_name)
                    else:
                        flex = find_column_flexible(df, [col_name])
                        if flex:
                            extra_found.append(flex)
                if extra_found:
                    extra_info = df[[sku_col] + extra_found].copy()
                    extra_info = extra_info.rename(columns={sku_col: SKU_COLUMN})
                    extra_info = extra_info.drop_duplicates(subset=[SKU_COLUMN])
                    log(f"✅ تم استخراج {extra_found} من {filename}")

            log(f"✅ تحميل {filename}: {len(temp)} SKU")

        if not sales_data:
            return None

        sales_combined = None
        for year, data in sales_data.items():
            if sales_combined is None:
                sales_combined = data
            else:
                sales_combined = pd.merge(sales_combined, data, on=SKU_COLUMN, how="outer")

        if extra_info is not None:
            sales_combined = pd.merge(sales_combined, extra_info, on=SKU_COLUMN, how="left")
            for col in SALES_EXTRA_COLUMNS:
                if col not in sales_combined.columns:
                    sales_combined[col] = ""

        # عرض عينة من قيم SKU في المبيعات للتحقق
        sample_skus = sales_combined[SKU_COLUMN].head(5).tolist()
        log(f"📊 عينة من SKU في المبيعات: {sample_skus}")

        return sales_combined

    except Exception as e:
        log(f"❌ خطأ في تحميل المبيعات: {e}")
        return None

def load_cogs_data():
    """
    قراءة ملف التكلفة (COGS) بشكل مرن جدًا مع تطبيع SKU.
    يبحث عن عمود SKU وأعمدة السعر للسنوات 2025 و 2026 بأي صيغة ممكنة.
    """
    try:
        if not os.path.exists(COGS_FILE_PATH):
            log(f"⚠️ ملف COGS غير موجود: {COGS_FILE_PATH}")
            return None

        df = pd.read_excel(COGS_FILE_PATH)
        # تنظيف أسماء الأعمدة
        df = clean_column_names(df)
        log(f"📄 أعمدة COGS: {list(df.columns)}")

        # البحث عن عمود SKU
        sku_col = find_column_flexible(df, ["SKU", "ARTCODE", "ITEM", "CODE"])
        if sku_col is None:
            log(f"❌ لم يتم العثور على عمود SKU في ملف COGS")
            return None

        # تطبيع SKU
        df = normalize_sku(df, sku_col)

        # البحث عن عمود السعر لعام 2025 و 2026 (بمرونة عالية)
        price_2025_col = None
        price_2026_col = None

        # قائمة موسعة من الأنماط الممكنة
        patterns_2025 = [
            "PRICE_PER_UNITE 25", "PRICE_PER_UNITE25", "PRICE PER UNIT 25", "PRICE_PER_UNITE_25",
            "PRICE_PER_UNITE2025", "PRICE PER UNIT 2025", "PRICE 2025", "UNIT PRICE 2025",
            "COST 2025", "PRICE_PER_UNITE_2025", "PRICE_PER_UNITE 2025"
        ]
        patterns_2026 = [
            "PRICE_PER_UNITE 26", "PRICE_PER_UNITE26", "PRICE PER UNIT 26", "PRICE_PER_UNITE_26",
            "PRICE_PER_UNITE2026", "PRICE PER UNIT 2026", "PRICE 2026", "UNIT PRICE 2026",
            "COST 2026", "PRICE_PER_UNITE_2026", "PRICE_PER_UNITE 2026"
        ]

        price_2025_col = find_column_flexible(df, patterns_2025)
        price_2026_col = find_column_flexible(df, patterns_2026)

        # إذا لم نجد، نبحث بأسلوب مختلف: أي عمود يحتوي على "price" ورقم السنة
        if not price_2025_col:
            for col in df.columns:
                col_lower = col.lower()
                if "price" in col_lower and ("25" in col_lower or "2025" in col_lower):
                    price_2025_col = col
                    break
        if not price_2026_col:
            for col in df.columns:
                col_lower = col.lower()
                if "price" in col_lower and ("26" in col_lower or "2026" in col_lower):
                    price_2026_col = col
                    break

        if not price_2025_col:
            log(f"❌ لم يتم العثور على عمود السعر لعام 2025 بعد كل المحاولات")
            return None
        if not price_2026_col:
            log(f"❌ لم يتم العثور على عمود السعر لعام 2026 بعد كل المحاولات")
            return None

        log(f"✅ في COGS: SKU='{sku_col}', 2025='{price_2025_col}', 2026='{price_2026_col}'")

        cogs_df = df[[sku_col, price_2025_col, price_2026_col]].copy()
        cogs_df = cogs_df.rename(columns={
            sku_col: SKU_COLUMN,
            price_2025_col: "Cost_2025",
            price_2026_col: "Cost_2026"
        })
        cogs_df["Cost_2025"] = pd.to_numeric(cogs_df["Cost_2025"], errors="coerce")
        cogs_df["Cost_2026"] = pd.to_numeric(cogs_df["Cost_2026"], errors="coerce")

        log(f"✅ تم تحميل بيانات COGS: {len(cogs_df)} SKU")

        # عرض عينة من قيم SKU في COGS للتحقق
        sample_skus = cogs_df[SKU_COLUMN].head(5).tolist()
        log(f"📊 عينة من SKU في COGS: {sample_skus}")

        return cogs_df

    except Exception as e:
        log(f"❌ خطأ في تحميل COGS: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════
#  المقارنة الأساسية
# ═══════════════════════════════════════════════════════════════════════════

def compare_stocks(old_file: str, new_file: str):
    df_old = pd.read_excel(old_file)
    df_new = pd.read_excel(new_file)

    # تنظيف أسماء الأعمدة
    df_old = clean_column_names(df_old)
    df_new = clean_column_names(df_new)

    if STOCK_COLUMN not in df_old.columns or STOCK_COLUMN not in df_new.columns:
        raise ValueError(f"❌ العمود '{STOCK_COLUMN}' غير موجود")
    if SKU_COLUMN not in df_old.columns or SKU_COLUMN not in df_new.columns:
        raise ValueError(f"❌ العمود '{SKU_COLUMN}' غير موجود")

    # تطبيع SKU في الملفين
    df_old = normalize_sku(df_old, SKU_COLUMN)
    df_new = normalize_sku(df_new, SKU_COLUMN)

    # تجميع الكميات
    df_old_grouped = df_old.groupby(SKU_COLUMN, as_index=False)[STOCK_COLUMN].sum()
    df_new_grouped = df_new.groupby(SKU_COLUMN, as_index=False)[STOCK_COLUMN].sum()

    # دمج البيانات
    merged = pd.merge(
        df_old_grouped.rename(columns={STOCK_COLUMN: "old_stock"}),
        df_new_grouped.rename(columns={STOCK_COLUMN: "new_stock"}),
        on=SKU_COLUMN, how="inner"
    )

    merged["old_stock"] = pd.to_numeric(merged["old_stock"], errors="coerce")
    merged["new_stock"] = pd.to_numeric(merged["new_stock"], errors="coerce")
    merged.dropna(subset=["old_stock", "new_stock"], inplace=True)
    merged = merged[merged["old_stock"] != 0]

    merged["change_pct"] = ((merged["new_stock"] - merged["old_stock"]) / merged["old_stock"].abs()) * 100
    merged["change_pct"] = merged["change_pct"].round(2)
    merged["absolute_drop"] = merged["old_stock"] - merged["new_stock"]

    # شرط Spikes
    condition = (merged["change_pct"] <= SPIKE_THRESHOLD) | (merged["absolute_drop"] >= MIN_ABSOLUTE_DROP)
    spikes = merged[condition].copy()

    # إضافة الأعمدة الإضافية من الملف الأحدث (OUTSTANDING و PR QTY)
    all_extra_cols = OUTSTANDING_COLS + OTHER_EXTRA_COLUMNS
    # تصفية الأعمدة الموجودة فعلاً في df_new لتجنب KeyError
    existing_extra = [col for col in all_extra_cols if col in df_new.columns]
    if existing_extra:
        extra_cols_df = df_new.groupby(SKU_COLUMN)[existing_extra].first().reset_index()
        spikes = spikes.merge(extra_cols_df, on=SKU_COLUMN, how="left")
    else:
        log("⚠️ لم يتم العثور على أعمدة OUTSTANDING أو PR QTY في الملف الأحدث")

    # جمع الـ Outstanding
    for col in OUTSTANDING_COLS:
        if col not in spikes.columns:
            spikes[col] = 0
        spikes[col] = spikes[col].fillna(0)
    spikes["OUTSTANDING_TOTAL"] = spikes[OUTSTANDING_COLS[0]] + spikes[OUTSTANDING_COLS[1]] + spikes[OUTSTANDING_COLS[2]]

    if "PR QTY" not in spikes.columns:
        spikes["PR QTY"] = 0
    spikes["PR QTY"] = spikes["PR QTY"].fillna(0)

    spikes["zero_stock_alert"] = spikes["new_stock"].apply(lambda x: "⚠️ ZERO STOCK!" if x == 0 else "")

    # دمج بيانات المبيعات
    sales_df = load_sales_data()
    if sales_df is not None:
        spikes = spikes.merge(sales_df, on=SKU_COLUMN, how="left")
        spikes["Sales_2025"] = spikes["Sales_2025"].fillna(0)
        spikes["Sales_2026"] = spikes["Sales_2026"].fillna(0)
        for col in SALES_EXTRA_COLUMNS:
            if col not in spikes.columns:
                spikes[col] = ""
            else:
                spikes[col] = spikes[col].fillna("")
    else:
        spikes["Sales_2025"] = 0
        spikes["Sales_2026"] = 0
        for col in SALES_EXTRA_COLUMNS:
            spikes[col] = ""

    # دمج بيانات COGS (التكلفة)
    cogs_df = load_cogs_data()
    if cogs_df is not None:
        # تأكد من تطبيع SKU في cogs_df أيضاً
        cogs_df = normalize_sku(cogs_df, SKU_COLUMN)
        spikes = spikes.merge(cogs_df, on=SKU_COLUMN, how="left")
        spikes["Cost_2025"] = spikes["Cost_2025"].fillna(0)
        spikes["Cost_2026"] = spikes["Cost_2026"].fillna(0)
        matched = spikes[spikes["Cost_2025"] > 0].shape[0]
        log(f"✅ تم ربط {matched} من {len(spikes)} SKU مع بيانات COGS (Cost > 0)")
    else:
        spikes["Cost_2025"] = 0
        spikes["Cost_2026"] = 0

    spikes.sort_values("change_pct", inplace=True)
    spikes.reset_index(drop=True, inplace=True)

    return spikes

# ═══════════════════════════════════════════════════════════════════════════
#  حفظ التقرير (Excel)
# ═══════════════════════════════════════════════════════════════════════════

def save_report(spikes: pd.DataFrame, old_file: str, new_file: str) -> str:
    os.makedirs(REPORT_FOLDER, exist_ok=True)
    today_str = datetime.now().strftime("%d_%m_%Y")
    report_path = os.path.join(REPORT_FOLDER, f"spike_report_{today_str}.xlsx")

    date_old = extract_date_label(old_file)
    date_new = extract_date_label(new_file)

    # ترتيب الأعمدة
    column_order = [
        SKU_COLUMN,
        "CATEGORY",
        "Section_Ro_Name",
        "old_stock",
        "new_stock",
        "change_pct",
        "absolute_drop",
        "zero_stock_alert",
        "OUTSTANDING_TOTAL",
        "PR QTY",
        "Sales_2025",
        "Sales_2026",
        "Cost_2025",
        "Cost_2026"
    ]

    for col in column_order:
        if col not in spikes.columns:
            if col.startswith(("CATEGORY", "Section_Ro_Name", "zero_stock_alert")):
                spikes[col] = ""
            else:
                spikes[col] = 0

    available = [col for col in column_order if col in spikes.columns]
    export_df = spikes[available].rename(columns={
        SKU_COLUMN: "Item / SKU",
        "CATEGORY": "Category",
        "Section_Ro_Name": "Section Name",
        "old_stock": f"Stock {date_old}",
        "new_stock": f"Stock {date_new}",
        "change_pct": "Change %",
        "absolute_drop": "Drop (units)",
        "zero_stock_alert": "Zero Stock Alert",
        "OUTSTANDING_TOTAL": "Outstanding Total",
        "PR QTY": "PR QTY",
        "Sales_2025": "Sales 2025",
        "Sales_2026": "Sales 2026",
        "Cost_2025": "Cost per unit (SAR) 2025",
        "Cost_2026": "Cost per unit (SAR) 2026"
    })

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Spikes", index=False, startrow=2)
        wb = writer.book
        ws = writer.sheets["Spikes"]

        ws["A1"] = f"🚨 Daily Stock Spikes Report  |  {date_old}  →  {date_new}  |  Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True, size=12, color="C00000")
        ws.merge_cells(f"A1:{get_column_letter(len(export_df.columns))}1")

        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        for cell in ws[3]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align

        red_fill = PatternFill("solid", fgColor="FFD6D6")
        amber_fill = PatternFill("solid", fgColor="FFF3CD")
        zero_fill = PatternFill("solid", fgColor="B91C1C")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        pct_col_idx = export_df.columns.get_loc("Change %") + 1
        alert_col_idx = export_df.columns.get_loc("Zero Stock Alert") + 1

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            pct_val = None
            is_zero = False
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if cell.column == pct_col_idx:
                    try:
                        pct_val = float(cell.value)
                    except:
                        pass
                if cell.column == alert_col_idx and cell.value and "ZERO" in str(cell.value):
                    is_zero = True

            if is_zero:
                for cell in row:
                    cell.fill = zero_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                if pct_val is not None:
                    row[pct_col_idx-1].font = Font(color="FFFFFF", bold=True)
                    row[pct_col_idx-1].value = f"{pct_val:.1f}%"
            else:
                fill = red_fill if (pct_val is not None and pct_val <= -50) else amber_fill
                for cell in row:
                    cell.fill = fill
                if pct_val is not None:
                    pct_cell = row[pct_col_idx-1]
                    pct_cell.font = Font(bold=True, color="C00000" if pct_val <= -50 else "B45309")
                    pct_cell.value = f"{pct_val:.1f}%"

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    log(f"✅ الريبورت اتحفظ → {report_path}")
    return report_path

# ═══════════════════════════════════════════════════════════════════════════
#  بناء جدول HTML
# ═══════════════════════════════════════════════════════════════════════════

def build_html_table(spikes: pd.DataFrame, date_old: str, date_new: str) -> str:
    if spikes.empty:
        return "<p style='color:#16a34a;font-weight:bold;'>✅ No spikes detected today</p>"

    required_cols = [
        SKU_COLUMN,
        "CATEGORY",
        "Section_Ro_Name",
        "old_stock",
        "new_stock",
        "change_pct",
        "absolute_drop",
        "zero_stock_alert",
        "OUTSTANDING_TOTAL",
        "PR QTY",
        "Sales_2025",
        "Sales_2026",
        "Cost_2025",
        "Cost_2026"
    ]
    for col in required_cols:
        if col not in spikes.columns:
            spikes[col] = 0 if col.startswith(("old_", "new_", "change_", "absolute_", "Sales_", "Cost_", "OUTSTANDING", "PR")) else ""

    headers = [
        "Item / SKU", "Category", "Section Name",
        f"Stock ({date_old})", f"Stock ({date_new})", "Change %", "Drop (units)",
        "Alert", "Outstanding Total", "PR QTY",
        "Sales 2025", "Sales 2026", "Cost 2025 (SAR/unit)", "Cost 2026 (SAR/unit)"
    ]

    html = '<table style="border-collapse:collapse;width:100%;font-family:Segoe UI;font-size:13px;">'
    html += '<thead><tr style="background:#1f2937;color:#f9fafb;">'
    for h in headers:
        html += f'<th style="padding:11px 14px;border:1px solid #374151;text-align:center;">{h}</th>'
    html += ' </thead><tbody>'

    for _, row in spikes.iterrows():
        pct = row["change_pct"]
        zero_alert = row["zero_stock_alert"]
        pct_color = "#dc2626" if pct <= -50 else "#d97706"
        bg_color = "#fff1f2" if pct <= -50 else "#fffbeb"
        if zero_alert:
            bg_color = "#b91c1c"
            pct_color = "#ffffff"

        html += f'<tr style="background:{bg_color}; {"color:#ffffff;" if zero_alert else ""}">'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;font-weight:600;">{row[SKU_COLUMN]}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["CATEGORY"] if row["CATEGORY"] else "-"}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["Section_Ro_Name"] if row["Section_Ro_Name"] else "-"}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:right;">{row["old_stock"]:,.1f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:right;">{row["new_stock"]:,.1f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;color:{pct_color};font-weight:700;">{pct:.1f}%</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:right;">{row["absolute_drop"]:,.1f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;font-weight:bold;{"color:#ffffff;" if zero_alert else "color:#b91c1c;"}">{zero_alert}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["OUTSTANDING_TOTAL"]:,.1f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["PR QTY"]:,.1f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["Sales_2025"]:,.0f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["Sales_2026"]:,.0f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["Cost_2025"]:,.2f}</td>'
        html += f'<td style="padding:9px 14px;border:1px solid #e5e7eb;text-align:center;">{row["Cost_2026"]:,.2f}</td>'
        html += ' </tr>'

    html += '</tbody></table>'
    return html

# ═══════════════════════════════════════════════════════════════════════════
#  إرسال الإيميل
# ═══════════════════════════════════════════════════════════════════════════

def send_email(report_path: str, spikes: pd.DataFrame, old_file: str, new_file: str):
    date_old = extract_date_label(old_file)
    date_new = extract_date_label(new_file)
    today_str = datetime.now().strftime("%d/%m/%Y")
    count = len(spikes)

    subject = f"🚨 Stock Spikes Report — {today_str}  |  {count} item(s) flagged" if count > 0 else f"✅ Stock Report — {today_str}  |  No spikes detected"

    alert_box = ""
    if count > 0:
        alert_box = f'''
        <div style="background:#fef2f2;border-left:5px solid #dc2626;padding:14px 18px;margin:18px 0;border-radius:4px;">
          <span style="color:#dc2626;font-weight:700;">⚠️ {count} item(s) dropped > {abs(SPIKE_THRESHOLD):.0f}% OR ≥ {MIN_ABSOLUTE_DROP} units</span>
        </div>'''

    table_html = build_html_table(spikes, date_old, date_new)

    html_body = f'''
    <div style="font-family:Segoe UI;max-width:100%;">
      <div style="background:#1f2937;padding:20px 28px;border-radius:8px 8px 0 0;">
        <h2 style="margin:0;color:#f9fafb;">🔔 Daily Stock Spikes Report</h2>
        <p style="margin:6px 0 0;color:#9ca3af;">Comparing <b>{date_old}</b> → <b>{date_new}</b> &nbsp;|&nbsp; Rule: drop &gt; {abs(SPIKE_THRESHOLD):.0f}% OR drop ≥ {MIN_ABSOLUTE_DROP} units</p>
      </div>
      <div style="border:1px solid #e5e7eb;border-top:none;padding:22px 28px;border-radius:0 0 8px 8px;background:#fff; overflow-x:auto;">
        {alert_box}
        {table_html}
        <p style="margin-top:22px;color:#6b7280;font-size:11px;">📎 Full Excel report attached | Generated by <b>Spike Alarm</b></p>
      </div>
    </div>'''

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = EMAIL_TO
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Attachments.Add(report_path)
        mail.Send()
        log(f"✅ الإيميل اتبعت → {EMAIL_TO}")
    except Exception as e:
        log(f"❌ فشل إرسال الإيميل: {e}")

def send_error_email(error_msg: str):
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = EMAIL_TO
        mail.Subject = f"❌ Spike Alarm ERROR — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        mail.Body = f"خطأ في السكريبت:\n\n{error_msg}"
        mail.Send()
    except:
        pass

# ═══════════════════════════════════════════════════════════════════════════
#  التشغيل الرئيسي
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═" * 60)
    print(f"  🔔  SPIKE ALARM  —  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}")
    print("═" * 60)
    try:
        log("🔍 جاري البحث عن ملفات الإكسيل...")
        old_file, new_file = find_latest_two_files(EXCEL_FOLDER)

        log("📊 جاري المقارنة...")
        spikes = compare_stocks(old_file, new_file)
        log(f"    عدد الـ Spikes: {len(spikes)}")

        if not spikes.empty:
            print("\n📋 أول 20 صف من النتائج:")
            display_cols = [SKU_COLUMN, "old_stock", "new_stock", "change_pct", "absolute_drop", "zero_stock_alert", "OUTSTANDING_TOTAL", "Cost_2025", "Cost_2026"]
            print(spikes[display_cols].head(20).to_string(index=False))
            print()

        log("💾 جاري حفظ ريبورت Excel...")
        report_path = save_report(spikes, old_file, new_file)

        log("📧 جاري إرسال الإيميل عبر Outlook...")
        send_email(report_path, spikes, old_file, new_file)

        print("\n" + "═" * 60)
        print("  ✅  تم بنجاح!")
        print("═" * 60 + "\n")

    except Exception as e:
        err = traceback.format_exc()
        print(f"\n❌ خطأ:\n{err}")
        send_error_email(err)
        sys.exit(1)

if __name__ == "__main__":
    main()
