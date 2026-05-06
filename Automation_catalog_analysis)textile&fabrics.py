"""
Catalogus Layout Report Generator  (MERGED)
============================================
STEP 1 — Scan D:\\catalogus\\SCRIPTS\\original_layouts\\
          Read each .xlsx, find the calculation sheet, extract SKU + R.MTR
          Ask user to confirm / rename each sheet → saves layouts.xlsx

STEP 2 — Run the full report using layouts.xlsx + Helper files
          Output: catalogus_report.xlsx  (one sheet per layout)

Script Location  : D:\\catalogus\\SCRIPTS\\catalogus_report.py
Helper Files Dir : D:\\catalogus\\SCRIPTS\\Helpers\\
  - daily_stk.xlsx
  - monthly_demand.xlsx
  - sales.xlsx
Yearly Sales Dir : D:\\mahmoud_analysis_dashboard_folder\\spikes\\sales_files\\
  - 2024sales.xlsx / 2025sales.xlsx / 2026sales.xlsx
"""

import os
import re
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
ORIGINAL_LAYOUTS = os.path.join(SCRIPT_DIR, "original_layouts")
HELPERS_DIR      = os.path.join(SCRIPT_DIR, "Helpers")
YEARLY_SALES_DIR = r"D:\mahmoud_analysis_dashboard_folder\spikes\sales_files"

LAYOUTS_FILE = os.path.join(HELPERS_DIR, "layouts.xlsx")
DAILY_STK    = os.path.join(HELPERS_DIR, "daily_stk.xlsx")
MONTHLY_DMD  = os.path.join(HELPERS_DIR, "monthly_demand.xlsx")
SALES_FILE   = os.path.join(HELPERS_DIR, "sales.xlsx")

YEARS = [2024, 2025, 2026]

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_TEAL  = "FF1F5C6B"
MID_TEAL   = "FF2E7D8C"
LIGHT_TEAL = "FFD6EEF2"
GOLD       = "FFFFC000"
LIGHT_GOLD = "FFFFF2CC"
ORANGE_HDR = "FFED7D31"
LIGHT_ORG  = "FFFCE4D6"
WHITE      = "FFFFFFFF"
PURPLE_HDR = "FF7030A0"
LIGHT_PURP = "FFF2E7FF"

THIN   = Side(style="thin", color="FFB0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Style helpers ─────────────────────────────────────────────────────────────
def solid(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def style_hdr(cell, bg, fg="FFFFFFFF", sz=10):
    cell.font      = Font(name="Arial", bold=True, size=sz, color=fg)
    cell.fill      = solid(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def style_data(cell, bg=None, bold=False, color="FF000000", left=False):
    cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left" if left else "center", vertical="center")
    cell.border    = BORDER
    if bg:
        cell.fill  = solid(bg)

def to_int(val):
    try:
        if pd.isna(val):
            return ""
        return int(round(float(val)))
    except Exception:
        return ""

# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Extract SKUs from original_layouts
# ═══════════════════════════════════════════════════════════════════════════════

def find_calc_sheet(wb):
    """
    Try to find the calculation sheet.
    1. Exact / fuzzy match for 'CAT.*CALC' pattern.
    2. If not found, show user a numbered list to pick from.
    Returns the worksheet or None.
    """
    sheet_names = wb.sheetnames

    # Fuzzy match: sheet name contains both 'CAT' and 'CALC' (handles typos)
    for name in sheet_names:
        upper = name.strip().upper()
        if "CAT" in upper and "CALC" in upper:
            return wb[name], name

    # Nothing matched — ask user
    print("\n  Could not auto-detect calculation sheet.")
    print("  Available sheets:")
    for i, name in enumerate(sheet_names, 1):
        print(f"    {i}. {name}")
    while True:
        choice = input("  Enter number (or 0 to skip this file): ").strip()
        if choice == "0":
            return None, None
        if choice.isdigit() and 1 <= int(choice) <= len(sheet_names):
            chosen = sheet_names[int(choice) - 1]
            return wb[chosen], chosen
        print("  Invalid input, try again.")


def find_sku_and_rmtr(ws):
    """
    SKU column is always column C (col 3).
    Data always starts at row 7 (C7), regardless of what C6 contains.
    R.MTR column is found by scanning row 6 for a header containing MTR.
    """
    SKU_COL   = 3   # column C
    START_ROW = 7   # always start reading SKUs from row 7

    # Find R.MTR column: scan row 6 first
    rmtr_col = None
    for cell in ws[6]:
        val = str(cell.value).strip().upper() if cell.value else ""
        if "MTR" in val:
            rmtr_col = cell.column
            break

    # Fallback: scan rows 1-10
    if not rmtr_col:
        for r in range(1, 11):
            for cell in ws[r]:
                val = str(cell.value).strip().upper() if cell.value else ""
                if "MTR" in val:
                    rmtr_col = cell.column
                    break
            if rmtr_col:
                break

    data = []
    for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
        sku_cell = row[SKU_COL - 1] if SKU_COL <= len(row) else None
        if not sku_cell:
            continue
        sku_val = str(sku_cell.value).strip() if sku_cell.value is not None else ""
        if not re.fullmatch(r'\d{10}', sku_val):
            if data:
                break      # SKUs exhausted
            continue
        rmtr_val = None
        if rmtr_col and rmtr_col <= len(row):
            rmtr_val = row[rmtr_col - 1].value
        data.append((sku_val, rmtr_val))

    return data


def sanitize(name):
    for ch in r'\/:*?[]':
        name = name.replace(ch, "")
    return name.strip()[:31]


def suggest_name(filename):
    name = os.path.splitext(filename)[0]
    match = re.search(r'(?i)catalog', name)
    if match:
        return sanitize(name[:match.start()])
    return sanitize(name)


def step1_extract_layouts():
    """
    Scan original_layouts folder, extract SKU+R.MTR, ask for sheet names,
    then write layouts.xlsx into Helpers folder.
    Returns dict of {sheet_name: DataFrame(SKU, R.MTR)} or None on failure.
    """
    sep = "=" * 62

    print(sep)
    print("  STEP 1 — EXTRACT SKUs FROM ORIGINAL LAYOUTS")
    print(sep)

    if not os.path.exists(ORIGINAL_LAYOUTS):
        print(f"\n  [ERROR] Folder not found:\n    {ORIGINAL_LAYOUTS}")
        return None

    xlsx_files = sorted([f for f in os.listdir(ORIGINAL_LAYOUTS)
                         if f.lower().endswith(".xlsx")])
    if not xlsx_files:
        print("  No .xlsx files found in original_layouts folder.")
        return None

    print(f"\n  Scanning {len(xlsx_files)} file(s)...\n")

    valid = []   # list of (filename, sheet_name_used, data_list)

    for filename in xlsx_files:
        filepath = os.path.join(ORIGINAL_LAYOUTS, filename)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"  [ERROR] {filename}: {e}")
            continue

        ws, used_sheet = find_calc_sheet(wb)
        if ws is None:
            print(f"  [SKIP]  {filename}")
            continue

        data = find_sku_and_rmtr(ws)
        if not data:
            print(f"  [SKIP]  {filename}  (no SKUs found in sheet '{used_sheet}')")
            continue

        print(f"  [OK]    {filename}  ({len(data)} SKUs  ←  sheet: '{used_sheet}')")
        valid.append((filename, data))

    if not valid:
        print("\n  No valid files to process.")
        return None

    # ── Ask for sheet names ───────────────────────────────────────────────────
    print(f"\n{sep}")
    print("  NAME YOUR LAYOUT SHEETS")
    print("  Max 31 chars | Press Enter to accept suggestion")
    print(sep + "\n")

    sheet_map = {}   # filename -> chosen sheet name
    used_names = set()

    for filename, data in valid:
        suggested = suggest_name(filename)
        base = suggested
        c = 1
        while suggested.upper() in used_names:
            suggested = f"{base[:28]}_{c}"
            c += 1

        print(f"  File    : {filename}")
        print(f"  SKUs    : {len(data)}")
        print(f"  Default : [{suggested}]")
        user_input = input("  Name    : ").strip()

        chosen = sanitize(user_input) if user_input else suggested
        if not chosen:
            chosen = suggested

        base = chosen
        c = 1
        while chosen.upper() in used_names:
            chosen = f"{base[:28]}_{c}"
            c += 1
            print(f"  (Duplicate → renamed to: {chosen})")

        used_names.add(chosen.upper())
        sheet_map[filename] = chosen
        print(f"  → '{chosen}'\n")

    # ── Confirm ───────────────────────────────────────────────────────────────
    print(sep)
    print("  LAYOUT SHEETS TO CREATE:")
    for fn, sn in sheet_map.items():
        cnt = len(next(d for f, d in valid if f == fn))
        print(f"    '{sn}'  ({cnt} SKUs)")
    print(f"\n  Will save to: {LAYOUTS_FILE}")
    print(sep)
    go = input("\n  Proceed? [Enter=Yes / n=Cancel]: ").strip().lower()
    if go == "n":
        print("  Cancelled.")
        return None

    # ── Write layouts.xlsx ────────────────────────────────────────────────────
    os.makedirs(HELPERS_DIR, exist_ok=True)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    result_dict = {}

    for filename, data in valid:
        sname = sheet_map[filename]
        ws = out_wb.create_sheet(title=sname)

        # Header
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 15
        ws.row_dimensions[1].height = 22
        for col, header in enumerate(["SKU", "R.MTR"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = PatternFill("solid", start_color="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for i, (sku, rmtr) in enumerate(data, start=2):
            ws.cell(row=i, column=1, value=sku)
            ws.cell(row=i, column=2, value=rmtr)
            shade = (i % 2 == 0)
            for col in [1, 2]:
                c = ws.cell(row=i, column=col)
                c.font      = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                if shade:
                    c.fill = PatternFill("solid", start_color="EEF2FF")

        result_dict[sname] = pd.DataFrame(data, columns=["SKU", "R.MTR"])

    out_wb.save(LAYOUTS_FILE)
    print(f"\n  layouts.xlsx saved → {LAYOUTS_FILE}")
    print(f"  Total layout sheets: {len(result_dict)}\n")
    return result_dict


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Full Report Generator
# ═══════════════════════════════════════════════════════════════════════════════

MAIN_COLS = [
    ("SKU",                      14),
    ("R.MTR",                     8),
    ("STOCK WITHOUT SMALL ROLL", 16),
    ("NET STK",                  10),
    ("avg_monthly_demand",       16),
    ("COVERAGE MONTHS",          14),
    ("OUTS_TOTAL",               12),
    ("PR QTY",                    9),
    ("Remark",                   14),
]
YEARLY_COLS    = [(f"Sales_{y}", 12) for y in YEARS]
ALL_MAIN_COLS  = MAIN_COLS + YEARLY_COLS
INT_COLS       = {"R.MTR", "STOCK WITHOUT SMALL ROLL", "NET STK",
                  "avg_monthly_demand", "OUTS_TOTAL", "PR QTY",
                  "Sales_2024", "Sales_2025", "Sales_2026"}


def load_helpers():
    print("  Loading helper files …")

    daily = pd.read_excel(DAILY_STK, dtype={"ARTCODE": str, "ARTSEC": str})
    daily.columns = daily.columns.str.strip()
    daily["ARTCODE"] = daily["ARTCODE"].str.strip()
    daily["ARTSEC"]  = daily["ARTSEC"].str.strip()

    demand = pd.read_excel(MONTHLY_DMD, dtype={"SKU": str})
    demand.columns = demand.columns.str.strip()
    demand["SKU"] = demand["SKU"].str.strip()

    sales = pd.read_excel(SALES_FILE, dtype={"SKU": str})
    sales.columns = sales.columns.str.strip()
    sales["SKU"] = sales["SKU"].str.strip()
    sales["total_with_old"] = pd.to_numeric(sales["total_with_old"], errors="coerce")

    return daily, demand, sales


def load_yearly_sales():
    yearly = {}
    for year in YEARS:
        path = os.path.join(YEARLY_SALES_DIR, f"{year}sales.xlsx")
        if not os.path.exists(path):
            print(f"  [WARN] Not found, skipping: {path}")
            continue
        df = pd.read_excel(path, dtype={"SKU": str})
        df.columns = df.columns.map(lambda x: str(x).strip())
        df["SKU"] = df["SKU"].str.strip()
        total_col = [c for c in df.columns if c.upper() == "TOTAL"]
        if not total_col:
            print(f"  [WARN] No TOTAL column in {year}sales.xlsx — skipping.")
            continue
        df = df[["SKU", total_col[0]]].rename(columns={total_col[0]: f"Sales_{year}"})
        df[f"Sales_{year}"] = pd.to_numeric(df[f"Sales_{year}"], errors="coerce")
        yearly[year] = df
        print(f"  Loaded {year}sales.xlsx → {len(df):,} rows")
    return yearly


def build_layout_df(layout_df, daily, demand, sales_df, yearly_sales):
    df = layout_df.copy()
    df.columns  = df.columns.str.strip()
    df["SKU"]   = df["SKU"].astype(str).str.strip()
    df["R.MTR"] = pd.to_numeric(df["R.MTR"], errors="coerce").fillna(0)

    daily_sub = daily[["ARTCODE", "STOCK WITHOUT SMALL ROLL",
                        "OUTS_TOTAL", "PR QTY"]].copy()
    daily_sub = daily_sub.rename(columns={"ARTCODE": "SKU"})
    for c in ["STOCK WITHOUT SMALL ROLL", "OUTS_TOTAL", "PR QTY"]:
        daily_sub[c] = pd.to_numeric(daily_sub[c], errors="coerce").fillna(0)

    merged = df.merge(daily_sub, on="SKU", how="left")
    merged["STOCK WITHOUT SMALL ROLL"] = merged["STOCK WITHOUT SMALL ROLL"].fillna(0)
    merged["NET STK"] = merged["STOCK WITHOUT SMALL ROLL"] - merged["R.MTR"]

    demand_sub = demand[["SKU", "avg_monthly_demand"]].copy()
    demand_sub["avg_monthly_demand"] = pd.to_numeric(
        demand_sub["avg_monthly_demand"], errors="coerce")
    merged = merged.merge(demand_sub, on="SKU", how="left")

    def coverage(row):
        avg = row["avg_monthly_demand"]
        if pd.isna(avg) or avg == 0:
            return None
        return round(row["NET STK"] / avg, 2)
    merged["COVERAGE MONTHS"] = merged.apply(coverage, axis=1)

    def remark(row):
        net  = row["NET STK"]
        pr   = 0 if pd.isna(row.get("PR QTY",    0)) else row.get("PR QTY",    0)
        outs = 0 if pd.isna(row.get("OUTS_TOTAL", 0)) else row.get("OUTS_TOTAL", 0)
        cov  = row["COVERAGE MONTHS"]
        if net <= 0:
            return "Alternative"
        if net < 100 and pr == 0 and outs == 0:
            return "Repeat"
        if cov is not None and cov <= 6 and pr == 0 and outs == 0:
            return "Repeat"
        return ""
    merged["Remark"] = merged.apply(remark, axis=1)

    sales_sub = sales_df[["SKU", "total_with_old"]].copy()
    merged = merged.merge(sales_sub, on="SKU", how="left")

    for year in YEARS:
        col = f"Sales_{year}"
        if year in yearly_sales:
            merged = merged.merge(yearly_sales[year], on="SKU", how="left")
        else:
            merged[col] = None

    return merged


def write_main_table(ws, df):
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(MAIN_COLS))
    t = ws.cell(row=1, column=1, value="Layout Stock Analysis")
    t.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t.fill      = solid(DARK_TEAL)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ystart = len(MAIN_COLS) + 1
    yend   = len(ALL_MAIN_COLS)
    ws.merge_cells(start_row=1, start_column=ystart, end_row=1, end_column=yend)
    yt = ws.cell(row=1, column=ystart, value="Yearly Sales")
    yt.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    yt.fill      = solid(PURPLE_HDR)
    yt.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for ci, (col_name, col_w) in enumerate(ALL_MAIN_COLS, 1):
        is_yearly = col_name.startswith("Sales_")
        bg = PURPLE_HDR if is_yearly else MID_TEAL
        display = col_name.upper()
        if is_yearly:
            year = col_name.split("_")[1]
            display = f"SALES {year}"
        cell = ws.cell(row=2, column=ci, value=display)
        style_hdr(cell, bg)
        ws.column_dimensions[get_column_letter(ci)].width = col_w
    ws.row_dimensions[2].height = 32

    data_keys = [c[0] for c in ALL_MAIN_COLS]
    for ri, (_, row) in enumerate(df.iterrows(), 1):
        er = 2 + ri
        bg_base   = LIGHT_TEAL if ri % 2 == 0 else WHITE
        bg_yearly = LIGHT_PURP if ri % 2 == 0 else WHITE
        for ci, col_name in enumerate(data_keys, 1):
            is_yearly = col_name.startswith("Sales_")
            raw = row.get(col_name, "")
            val = to_int(raw) if col_name in INT_COLS else (
                "" if pd.isna(raw) else raw)
            cell = ws.cell(row=er, column=ci, value=val)
            if col_name == "Remark" and val:
                style_data(cell, bg="FFFFDEDE", bold=True, color="FFCC0000")
            elif is_yearly:
                style_data(cell, bg=bg_yearly)
            else:
                style_data(cell, bg=bg_base)
        ws.row_dimensions[er].height = 16


def write_top10(ws, section_df, sales_df, start_col):
    ws.merge_cells(start_row=1, start_column=start_col,
                   end_row=1,   end_column=start_col + 1)
    t = ws.cell(row=1, column=start_col, value="Main Collection Top 10")
    t.font      = Font(name="Arial", bold=True, size=10, color="FF000000")
    t.fill      = solid(GOLD)
    t.alignment = Alignment(horizontal="center", vertical="center")

    for i, hdr in enumerate(["SKU", "Sales"], 0):
        cell = ws.cell(row=2, column=start_col + i, value=hdr)
        style_hdr(cell, GOLD, fg="FF000000")
        ws.column_dimensions[get_column_letter(start_col + i)].width = 16

    if section_df is None or section_df.empty:
        return

    sec = section_df[["ARTCODE"]].copy()
    sec["ARTCODE"] = sec["ARTCODE"].astype(str).str.strip()
    sec = sec.merge(
        sales_df[["SKU", "total_with_old"]].rename(columns={"SKU": "ARTCODE"}),
        on="ARTCODE", how="left")
    sec["total_with_old"] = pd.to_numeric(
        sec["total_with_old"], errors="coerce").fillna(0)
    top10 = sec.sort_values("total_with_old", ascending=False).head(10).reset_index(drop=True)

    for ri, (_, row) in enumerate(top10.iterrows(), 1):
        er = 2 + ri
        bg = LIGHT_GOLD if ri % 2 == 0 else WHITE
        style_data(ws.cell(row=er, column=start_col,
                            value=str(row["ARTCODE"])), bg=bg)
        style_data(ws.cell(row=er, column=start_col + 1,
                            value=to_int(row["total_with_old"])), bg=bg)
        ws.row_dimensions[er].height = 16


def write_main_collection(ws, section_df, layout_skus, sales_df, start_col):
    mc_hdrs = [("Main Collection SKUS", 20),
               ("Total_Qty",            12),
               ("STK",                  10)]

    ws.merge_cells(start_row=1, start_column=start_col,
                   end_row=1,   end_column=start_col + len(mc_hdrs) - 1)
    t = ws.cell(row=1, column=start_col, value="Main Collection SKUS")
    t.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    t.fill      = solid(ORANGE_HDR)
    t.alignment = Alignment(horizontal="center", vertical="center")

    for i, (hdr, w) in enumerate(mc_hdrs):
        cell = ws.cell(row=2, column=start_col + i, value=hdr)
        style_hdr(cell, ORANGE_HDR)
        ws.column_dimensions[get_column_letter(start_col + i)].width = w

    ws.column_dimensions[
        get_column_letter(start_col + len(mc_hdrs))].width = 28

    if section_df is None or section_df.empty:
        return

    sec = section_df[["ARTCODE", "STOCK WITHOUT SMALL ROLL"]].copy()
    sec["ARTCODE"] = sec["ARTCODE"].astype(str).str.strip()
    sec["STOCK WITHOUT SMALL ROLL"] = pd.to_numeric(
        sec["STOCK WITHOUT SMALL ROLL"], errors="coerce").fillna(0)
    sec = sec.merge(
        sales_df[["SKU", "total_with_old"]].rename(columns={"SKU": "ARTCODE"}),
        on="ARTCODE", how="left")
    sec["total_with_old"] = pd.to_numeric(
        sec["total_with_old"], errors="coerce").fillna(0)
    sec = sec.sort_values("total_with_old", ascending=False).reset_index(drop=True)

    for ri, (_, row) in enumerate(sec.iterrows(), 1):
        er      = 2 + ri
        sku     = str(row["ARTCODE"])
        stk     = to_int(row["STOCK WITHOUT SMALL ROLL"])
        sal     = to_int(row["total_with_old"])
        bg      = LIGHT_ORG if ri % 2 == 0 else WHITE
        missing = sku not in layout_skus

        sku_cell = ws.cell(row=er, column=start_col, value=sku)
        if missing:
            style_data(sku_cell, bg="FFFFD7D7", color="FFCC0000", left=True)
            note = ws.cell(row=er, column=start_col + len(mc_hdrs),
                           value="lost from the main collection")
            note.font      = Font(name="Arial", size=8, color="FFCC0000", italic=True)
            note.alignment = Alignment(horizontal="left", vertical="center")
        else:
            style_data(sku_cell, bg=bg, left=True)

        style_data(ws.cell(row=er, column=start_col + 1, value=sal), bg=bg)
        style_data(ws.cell(row=er, column=start_col + 2, value=stk), bg=bg)
        ws.row_dimensions[er].height = 16


def step2_build_report(layouts_dict):
    sep = "=" * 62
    print(sep)
    print("  STEP 2 — BUILD FULL REPORT")
    print(sep + "\n")

    for f in [DAILY_STK, MONTHLY_DMD, SALES_FILE]:
        if not os.path.exists(f):
            print(f"  [ERROR] Helper file not found: {f}")
            sys.exit(1)

    if not os.path.isdir(YEARLY_SALES_DIR):
        print(f"  [WARN] Yearly sales folder not found: {YEARLY_SALES_DIR}")
        print("         Yearly sales columns will be empty.\n")

    daily, demand, sales_df = load_helpers()
    yearly_sales = load_yearly_sales()
    print()

    layout_names = list(layouts_dict.keys())

    print("─" * 55)
    print("  Enter the Main Section (ARTSEC) for each layout.")
    print("  Press ENTER to skip.\n")

    main_section_map = {}
    for name in layout_names:
        val = input(f"  Main section for '{name}': ").strip()
        main_section_map[name] = val if val else None
    print()

    top10_start = len(ALL_MAIN_COLS) + 2
    mc_start    = top10_start + 2 + 2

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, layout_df in layouts_dict.items():
        print(f"  Processing: {sheet_name} …")
        try:
            merged      = build_layout_df(layout_df, daily, demand,
                                          sales_df, yearly_sales)
            layout_skus = set(merged["SKU"].astype(str).str.strip())

            section_code = main_section_map.get(sheet_name)
            section_df   = (
                daily[daily["ARTSEC"] == str(section_code)].copy()
                if section_code else pd.DataFrame()
            )

            ws = wb.create_sheet(title=sheet_name[:31])
            ws.freeze_panes = "A3"

            write_main_table(ws, merged)
            write_top10(ws, section_df, sales_df, start_col=top10_start)
            write_main_collection(ws, section_df, layout_skus,
                                  sales_df, start_col=mc_start)

        except Exception as e:
            print(f"  [ERROR] Layout '{sheet_name}': {e}")
            import traceback; traceback.print_exc()

    out_path = os.path.join(SCRIPT_DIR, "catalogus_report.xlsx")
    wb.save(out_path)
    print(f"\n  Done!  Report saved → {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print("\n")
    print("╔══════════════════════════════════════════════════════════╗")
    print("║          CATALOGUS LAYOUT REPORT GENERATOR              ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    # STEP 1 — extract from original_layouts → layouts.xlsx
    layouts_dict = step1_extract_layouts()
    if not layouts_dict:
        print("\n  No layouts extracted. Exiting.")
        input("\n  Press Enter to exit...")
        return

    print()

    # STEP 2 — build the full report
    step2_build_report(layouts_dict)

    input("\n  Press Enter to exit...")


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as _err:
        print("\n")
        print("!" * 62)
        print("  UNEXPECTED ERROR — full details below:")
        print("!" * 62)
        traceback.print_exc()
        print("!" * 62)
        input("\n  Press Enter to exit...")
