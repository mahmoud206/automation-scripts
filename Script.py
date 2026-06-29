"""
Activation Daily Monitor
Schedule in Task Scheduler 3x per day (e.g. 08:00, 13:00, 17:00).
Each run fires arrival notifications when within +-3 days of Arriving Date.
No wrap_text, no merged cells anywhere in any output.
"""

import os, sys, glob, logging
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE          = r"D:\mahmoud_analysis_dashboard_folder\Activation"
STK_FOLDER    = os.path.join(BASE, "Daily STK")
ACT_FILE      = os.path.join(BASE, "Activated Collections", "Activated_sections.xlsx")
REPORT_FOLDER = os.path.join(BASE, "Daily Report")
os.makedirs(REPORT_FOLDER, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(BASE, "monitor.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

C_HEADER = "1F4E79"; C_RED = "FF0000"; C_ORANGE = "FF8C00"
C_GREEN  = "00B050"; C_YELLOW = "FFC000"; C_GREY = "D9D9D9"
C_WHITE  = "FFFFFF"; C_LGREEN = "E2EFDA"

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=False)
AC = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ── CSV ────────────────────────────────────────────────────────────────────────

def find_csv(d: date):
    name = f"daily stock {d.strftime('%d_%m_%Y')}.csv"
    p    = os.path.join(STK_FOLDER, name)
    if os.path.isfile(p): return p
    for f in glob.glob(os.path.join(STK_FOLDER, "daily stock *.csv")):
        if os.path.basename(f).lower() == name.lower(): return f
    return None

def load_csv(p: str) -> pd.DataFrame:
    df = pd.read_csv(p, dtype=str, encoding="utf-8-sig")
    df.columns = df.columns.str.strip()
    return df


# ── Stock helpers ──────────────────────────────────────────────────────────────

def _n(v) -> float:
    try:    return float(v)
    except: return 0.0

def _row(df, ac):
    m = df["ARTCODE"] == str(ac).strip()
    return df[m].iloc[0] if m.any() else None

def cur_stk(df, ac) -> float:
    r = _row(df, ac)
    if r is None: return 0.0
    return (_n(r.get("BARCODE STOCK",0)) + _n(r.get("OUTSTANDING",0))
            + _n(r.get("OUTSTANDING SP",0)) + _n(r.get("OUTSTANDING N",0)))

def pr_qty(df, ac) -> float:
    r = _row(df, ac)
    return _n(r.get("PR QTY",0)) if r is not None else 0.0

def out_qty(df, ac) -> float:
    r = _row(df, ac)
    if r is None: return 0.0
    return _n(r.get("OUTSTANDING",0)) + _n(r.get("OUTSTANDING SP",0)) + _n(r.get("OUTSTANDING N",0))


# ── Sheet helpers ──────────────────────────────────────────────────────────────

def find_hrow(ws):
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == "SKU": return cell.row
    return None

def col_idx(ws, hr, name):
    up = name.strip().upper()
    for c in ws[hr]:
        if str(c.value or "").strip().upper() == up: return c.column
    return None

def ensure_col(ws, hr, name) -> int:
    idx = col_idx(ws, hr, name)
    if idx is not None: return idx
    last = max((c for c in range(1, ws.max_column+1) if ws.cell(hr,c).value is not None), default=1)
    nc   = last + 1
    h    = ws.cell(hr, nc)
    h.value     = name
    h.font      = Font(bold=True, color=C_WHITE)
    h.fill      = PatternFill("solid", fgColor=C_HEADER)
    h.alignment = AC
    h.border    = BORDER
    ws.column_dimensions[get_column_letter(nc)].width = 32
    return nc

def read_date(ws, label):
    for row in ws.iter_rows():
        for c in row:
            if str(c.value or "").strip().upper() == label.upper():
                v = ws.cell(c.row, c.column+1).value
                if v is None: return None
                if isinstance(v, datetime): return v.date()
                if isinstance(v, date):     return v
                try:    return pd.to_datetime(str(v)).date()
                except: return None
    return None


# ── Notification ───────────────────────────────────────────────────────────────

def notify(title, msg):
    try:
        from win10toast import ToastNotifier
        ToastNotifier().show_toast(title, msg, duration=12, threaded=True)
    except Exception:
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, msg, title, 0x40)
        except Exception as e:
            log.warning(f"Notify failed: {e}")
    log.info(f"NOTIFY | {title} | {msg}")


# ── Process one collection sheet ───────────────────────────────────────────────

def process_sheet(ws, df_t, df_y, name) -> dict:
    hr = find_hrow(ws)
    if hr is None:
        log.warning(f"[{name}] No SKU header — skip")
        return {}

    c_sku  = col_idx(ws, hr, "SKU")
    c_hold = col_idx(ws, hr, "HOLDING STK")
    c_req  = col_idx(ws, hr, "REQ QTY")
    c_cur  = ensure_col(ws, hr, "CURRENT STK + OUT")
    c_stat = ensure_col(ws, hr, "STATUES")
    c_stks = ensure_col(ws, hr, "SKU STK STATUES")

    if c_sku is None:
        log.warning(f"[{name}] SKU col missing — skip")
        return {}

    today    = date.today()
    arr_date = read_date(ws, "ARRIVING DATE")

    rows_out = []; neg_list = []; urgent = []; pr_list = []; out_list = []

    for rn in range(hr+1, ws.max_row+1):
        ac = str(ws.cell(rn, c_sku).value or "").strip()
        if not ac or ac.upper() in ("NONE","NAN"): continue

        ct = cur_stk(df_t, ac)
        cy = cur_stk(df_y, ac)

        # CURRENT STK + OUT
        cc = ws.cell(rn, c_cur)
        cc.value = ct; cc.number_format = "#,##0"; cc.alignment = AC; cc.border = BORDER

        # STATUES
        prt  = pr_qty(df_t,  ac); pry  = pr_qty(df_y,  ac)
        outt = out_qty(df_t, ac); outy = out_qty(df_y, ac)

        sv = ""; sf = C_GREY
        if prt > 0:
            sv = "ASK JOSEPH ABOUT WHY PR IS PENDING?"; sf = C_YELLOW; pr_list.append(ac)
        elif pry > 0 and prt == 0 and outt > outy:
            sv = "ASK JOSEPH ABOUT THE SHIPPING DAY";   sf = C_ORANGE; out_list.append(ac)

        cs = ws.cell(rn, c_stat)
        cs.value = sv; cs.fill = PatternFill("solid", fgColor=sf)
        cs.font = Font(bold=bool(sv)); cs.alignment = AL; cs.border = BORDER

        # SKU STK STATUES
        req  = _n(ws.cell(rn, c_req).value)  if c_req  else 0.0
        hold = _n(ws.cell(rn, c_hold).value) if c_hold else 0.0
        pct  = (ct - cy) / cy * 100 if cy > 0 else 0.0

        parts = []
        if cy > 0:
            arrow = "\u25bc" if pct < 0 else "\u25b2"
            parts.append(f"{arrow} {abs(pct):.1f}% vs yesterday (Today: {ct:,.0f} | Yesterday: {cy:,.0f})")
        if hold > 0:
            dh = ct - hold
            dn = ""
            if arr_date:
                d = (arr_date - today).days
                dn = f" | Arriving in {d}d" if d>0 else (" | Arrives TODAY" if d==0 else f" | Arrived {abs(d)}d ago")
            parts.append(f"vs Holding: {dh:+,.0f}{dn}")
        if req > 0 and ct <= req:
            parts.append("REPEAT NOW — QTY INSUFFICIENT!")
            urgent.append(ac)

        sv2 = "  |  ".join(parts) if parts else "OK"
        csk = ws.cell(rn, c_stks)
        csk.value = sv2; csk.alignment = AL; csk.border = BORDER

        if "REPEAT NOW" in sv2:
            csk.fill = PatternFill("solid", fgColor=C_RED);    csk.font = Font(bold=True, color=C_WHITE)
        elif pct < -10:
            csk.fill = PatternFill("solid", fgColor=C_ORANGE); csk.font = Font(bold=True); neg_list.append(ac)
        else:
            csk.fill = PatternFill("solid", fgColor=C_LGREEN); csk.font = Font()

        rows_out.append(dict(ac=ac, ct=ct, cy=cy, pct=pct, req=req, hold=hold))

    n = len(rows_out)
    if urgent:        coll_st = "STUCKED"
    elif len(pr_list)==n and n>0: coll_st = "NOT COMPLETE"
    else:
        below  = sum(1 for r in rows_out if r["req"]>0 and r["ct"]<=r["req"])
        coll_st = "NOT COMPLETE" if below>0 else "COMPLETE"

    return dict(sheet_name=name, total=n, neg=len(neg_list), pr=pr_list,
                out=out_list, urgent=urgent, status=coll_st, arr=arr_date, rows=rows_out)


# ── Update Collections status ──────────────────────────────────────────────────

def update_collections(wb, summaries):
    ws = wb["Collections"]
    sc = nc = None
    for c in ws[1]:
        v = str(c.value or "").strip().upper()
        if v == "STATUES":       sc = c.column
        if "SECTION NAME" in v:  nc = c.column
    if sc is None: log.warning("Collections STATUES col not found"); return

    smap = {s["sheet_name"]: s["status"] for s in summaries}
    for rn in range(2, ws.max_row+1):
        sh = str(ws.cell(rn, nc if nc else 2).value or "").strip()
        if sh not in smap: continue
        st = smap[sh]; c = ws.cell(rn, sc)
        c.value = st; c.alignment = AC; c.border = BORDER
        if   st == "COMPLETE": c.fill = PatternFill("solid", fgColor=C_GREEN);  c.font = Font(bold=True, color=C_WHITE)
        elif st == "STUCKED":  c.fill = PatternFill("solid", fgColor=C_RED);    c.font = Font(bold=True, color=C_WHITE)
        else:                  c.fill = PatternFill("solid", fgColor=C_YELLOW); c.font = Font(bold=True)


# ── Build daily report ─────────────────────────────────────────────────────────

def build_report(summaries, today):
    wb = Workbook(); ws = wb.active; ws.title = "Daily Summary"
    hdrs   = ["Collection Name","SKU","# SKUs","Neg >10%","STATUS","Diff vs Holding","Arriving Date","Days to Arrival"]
    widths = [28, 16, 10, 12, 48, 20, 16, 16]
    for ci,(h,w) in enumerate(zip(hdrs,widths),1):
        c = ws.cell(1,ci)
        c.value=h; c.font=Font(bold=True,color=C_WHITE,size=10)
        c.fill=PatternFill("solid",fgColor=C_HEADER); c.alignment=AC; c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    rn = 2
    for s in summaries:
        parts=[]
        if s["pr"]:     parts.append(f"STILL PR ({len(s['pr'])} SKU)")
        if s["out"]:    parts.append(f"BECAME OUTSTANDING ({len(s['out'])} SKU)")
        if s["urgent"]: parts.append(f"URGENT!!! NEED REPEAT NOW ({len(s['urgent'])} SKU)")
        if not parts:   parts.append("OK")
        slbl = "  |  ".join(parts)

        arr = s["arr"]; days_str=""
        if arr:
            d = (arr-today).days
            days_str = f"In {d}d" if d>0 else ("TODAY" if d==0 else f"{abs(d)}d ago")

        if "URGENT" in slbl:      bg="FFD7D7"
        elif "OUTSTANDING" in slbl: bg="FFE8CC"
        elif "STILL PR" in slbl:    bg="FFFACD"
        else:                       bg=C_LGREEN

        for r in (s["rows"] if s["rows"] else [{}]):
            diff = f"{r.get('ct',0)-r.get('hold',0):+,.0f}" if r.get("hold",0)>0 else ""
            vals = [s["sheet_name"], r.get("ac",""), s["total"], s["neg"], slbl, diff,
                    arr.strftime("%d/%m/%Y") if arr else "", days_str]
            for ci,v in enumerate(vals,1):
                c=ws.cell(rn,ci); c.value=v
                c.fill=PatternFill("solid",fgColor=bg); c.border=BORDER
                c.alignment=AL; c.font=Font(size=9)
            rn += 1

    path = os.path.join(REPORT_FOLDER, f"{today.strftime('%d_%m_%Y')}.xlsx")
    wb.save(path); log.info(f"Report saved -> {path}"); return path


# ── Arrival notifications ──────────────────────────────────────────────────────

def check_notifications(summaries, today):
    for s in summaries:
        arr = s.get("arr")
        if arr is None: continue
        d = (arr - today).days
        if not (-3 <= d <= 3): continue
        nm = s["sheet_name"]
        if d > 0:   msg = f"[{nm}]  Arriving in {d} day(s)  —  {arr.strftime('%d/%m/%Y')}"
        elif d == 0:msg = f"[{nm}]  ARRIVING TODAY  —  {arr.strftime('%d/%m/%Y')}"
        else:        msg = f"[{nm}]  Expected {abs(d)} day(s) ago ({arr.strftime('%d/%m/%Y')}) — confirm receipt!"
        notify("Activation Alert", msg)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today = date.today(); yest = today - timedelta(days=1)
    log.info(f"=== Activation Monitor  {today} ===")

    pt = find_csv(today); py = find_csv(yest)
    if not pt:
        log.error(f"Today CSV missing: daily stock {today.strftime('%d_%m_%Y')}.csv"); sys.exit(1)
    if not py:
        log.warning("Yesterday CSV missing — PR comparison skipped")

    df_t = load_csv(pt)
    df_y = load_csv(py) if py else pd.DataFrame(columns=df_t.columns)

    if not os.path.isfile(ACT_FILE):
        log.error(f"Activated_sections.xlsx not found: {ACT_FILE}"); sys.exit(1)

    wb = load_workbook(ACT_FILE)
    collections = [str(r[1]).strip() for r in wb["Collections"].iter_rows(min_row=2, values_only=True) if r[1]]
    log.info(f"Collections: {collections}")

    summaries = []
    for sh in wb.sheetnames:
        if sh == "Collections" or sh not in collections: continue
        log.info(f"  Processing: {sh}")
        s = process_sheet(wb[sh], df_t, df_y, sh)
        if s: summaries.append(s)

    update_collections(wb, summaries)
    wb.save(ACT_FILE); log.info(f"Saved: {ACT_FILE}")

    if summaries: build_report(summaries, today)
    check_notifications(summaries, today)
    log.info("=== Done ===")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        # sys.exit() was called, we still want to pause
        pass
    except Exception as e:
        print("="*50)
        print("حدث خطأ غير متوقع:")
        import traceback
        traceback.print_exc()
        print("="*50)
    finally:
        input("\nاضغط Enter للخروج...")
